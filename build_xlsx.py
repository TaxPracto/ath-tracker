#!/usr/bin/env python3
"""Generate docs/ATH_Tracker_latest.xlsx from state (computed values, no formulas)."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))


def build(state):
    fins = state["finalists"]
    wb = Workbook()
    ws = wb.active
    ws.title = "ATH Screen"
    hdrs = ["Rank", "Symbol", "Company", "Board", "Sector Index", "Mkt Cap (Rs Cr)",
            "Entered ATH Zone", "Days in Zone", "Last Close", "ATH", "% from ATH", "ATH Date",
            "3M Return %", "6M Return %",
            "TTM Return %", "Nifty 500 %", "Nifty TM %", "Sector/SME %",
            "Alpha vs N500 pp", "Alpha vs Sector/SME pp",
            "TTM PAT (Rs Cr)", "Prior Peak PAT", "PAT vs Peak %", "Basis", "Reporting",
            "Latest Period", "New This Week"]
    ws.append(hdrs)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    new_fill = PatternFill("solid", fgColor="E2EFDA")
    thin = Border(bottom=Side(style="thin", color="D9D9D9"))
    for i, x in enumerate(fins, start=2):
        bench2 = x.get("sme_ret") if x["board"] == "SME" else x.get("sector_ret")
        sec = "SME Emerge" if x["board"] == "SME" else (x.get("sector_index") or "-")
        a1 = None if x.get("n500_ret") is None else round(x["stock_ret"] - x["n500_ret"], 1)
        a2 = None if bench2 is None else round(x["stock_ret"] - bench2, 1)
        patpk = (round((x["ttm_pat"] / x["prior_peak_pat"] - 1) * 100, 1)
                 if x.get("prior_peak_pat") and x["prior_peak_pat"] > 0 else None)
        board = ("SME-" + x.get("exch", "NSE")) if x["board"] == "SME" else "Main"
        row = [i - 1, x["symbol"], x["name"], board, sec, x.get("mcap"),
               x["zone_entry"], x["days_in_zone"], x["last_close"], x["ath"],
               x["pct_from_ath"], x["ath_date"], x.get("ret_3m"), x.get("ret_6m"),
               x["stock_ret"], x.get("n500_ret"),
               x.get("ntm_ret"), bench2, a1, a2, x.get("ttm_pat"), x.get("prior_peak_pat"),
               patpk, x.get("basis"), x.get("reporting"), x.get("latest_q"),
               "Yes" if x.get("is_new") else ("" if x.get("is_new") is None else "No")]
        for j, v in enumerate(row, start=1):
            ws.cell(i, j, v)
        if x.get("is_new"):
            for j in range(1, len(hdrs) + 1):
                ws.cell(i, j).fill = new_fill
    for row in ws.iter_rows(min_row=2, max_row=len(fins) + 1):
        for c in row:
            c.font = Font(name="Arial", size=10)
            c.border = thin
    widths = [5, 13, 32, 7, 20, 12, 12, 8, 10, 10, 9, 11, 9, 9, 9, 9, 9, 10, 9, 9, 11, 11, 9, 12, 11, 11, 9]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, j).column_letter].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:AA{len(fins)+1}"

    m = wb.create_sheet("Methodology")
    f = state["funnel"]
    for a, b in [
        ("ATH Radar - weekly screen", ""),
        ("Run date", state["run_date"] + " (prices as of " + state["asof"] + ")"),
        ("Universe", f"{f['mainboard']} NSE mainboard (EQ) + {f['sme']} NSE Emerge SME"),
        ("Filters", "Within 2% of all-time-high close; TTM return above Nifty 500, Nifty Total Market, sector index (mainboard) / NIFTY SME EMERGE (SME); TTM PAT >= every prior annual and rolling window (screener.in); mcap Rs 1,000-20,000 Cr"),
        ("Sort", "Newest entrant into the 2% ATH zone first; green rows new since previous run"),
        ("Funnel", f"{f['mainboard']+f['sme']} screened -> {f['ath_zone']} at ATH -> {f['outperforming']} outperforming -> {f['pat_at_ath']} record PAT -> {f['final']} in mcap band"),
        ("Caveats", "SME price history since Jul-2024, unadjusted for corporate actions; PAT history limited to screener-visible years; mainboard prices split-adjusted (Yahoo). Factual screen - not investment advice"),
    ]:
        m.append((a, b))
    for row in m.iter_rows():
        row[0].font = Font(name="Arial", bold=True, size=10)
        if len(row) > 1:
            row[1].font = Font(name="Arial", size=10)
            row[1].alignment = Alignment(wrap_text=True, vertical="top")
    m.column_dimensions["A"].width = 24
    m.column_dimensions["B"].width = 120

    out = os.path.join(BASE, "docs", "ATH_Tracker_latest.xlsx")
    wb.save(out)
    print("wrote", out)


if __name__ == "__main__":
    build(json.load(open(os.path.join(BASE, "data", "state.json"))))
